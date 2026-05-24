#!/usr/bin/env python3
"""
parse_input.py — normalize a list/spreadsheet into priority-board records.

Reads CSV / TSV / JSON (stdlib only) or Excel .xlsx (auto-installs openpyxl on
first use), maps messy column headers to a canonical schema, and prints:

    { "columns_detected": {<canonical>: <source header>, ...},
      "unmapped": [<source headers kept under each record's "extra">],
      "count": <int>,
      "records": [ {id,title,description,priority,status,estimate,labels,
                    assignee,due,project,url,extra}, ... ] }

Claude then assigns each record a lane (now/next/later/cut) + order + rationale
and feeds the result to build_board.py. See reference.md for the field aliases
and the prioritization rubric.

Usage:
    python3 parse_input.py --input tickets.csv
    python3 parse_input.py --input export.xlsx --sheet 0
    python3 parse_input.py --input list.tsv --format tsv --out records.json
"""
import argparse
import csv
import io
import json
import os
import re
import subprocess
import sys

# Canonical field -> list of accepted source-header aliases (lowercased, symbols stripped).
ALIASES = {
    "id":          ["id", "key", "identifier", "ticket", "ticketid", "issue",
                    "issueid", "issuekey", "number", "no", "ref"],
    "title":       ["title", "name", "summary", "issuetitle", "task", "item",
                    "headline", "subject", "story"],
    "description": ["description", "desc", "details", "detail", "body", "notes",
                    "note", "content"],
    "priority":    ["priority", "urgency", "prio", "p", "importance", "severity"],
    "status":      ["status", "state", "workflow", "workflowstate", "stage", "column"],
    "estimate":    ["estimate", "estimatepoints", "points", "point", "effort",
                    "sp", "storypoints", "size", "complexity"],
    "labels":      ["labels", "label", "tags", "tag", "categories", "category", "type"],
    "assignee":    ["assignee", "owner", "assignedto", "assigned", "responsible", "lead"],
    "due":         ["due", "duedate", "targetdate", "target", "deadline", "date", "eta"],
    "project":     ["project", "team", "cycle", "milestone", "epic", "sprint", "group"],
    "url":         ["url", "link", "href", "permalink", "weburl", "issueurl"],
}
CANONICAL = list(ALIASES.keys())


def _norm_header(h):
    """Lowercase a header and strip everything but a-z0-9 for fuzzy matching."""
    return re.sub(r"[^a-z0-9]", "", str(h or "").strip().lower())


def build_header_map(headers):
    """Return (mapping canonical->source_header, list of unmapped source headers)."""
    norm_to_alias = {}
    for canon, names in ALIASES.items():
        for n in names:
            norm_to_alias[_norm_header(n)] = canon

    mapping, used_sources = {}, set()
    # First pass: exact normalized match.
    for h in headers:
        nh = _norm_header(h)
        canon = norm_to_alias.get(nh)
        if canon and canon not in mapping:
            mapping[canon] = h
            used_sources.add(h)
    # Second pass: substring match for anything still unmapped (e.g. "issue title").
    for h in headers:
        if h in used_sources:
            continue
        nh = _norm_header(h)
        if not nh:
            continue
        for alias_norm, canon in norm_to_alias.items():
            if canon in mapping:
                continue
            if len(alias_norm) >= 3 and (alias_norm in nh or nh in alias_norm):
                mapping[canon] = h
                used_sources.add(h)
                break
    unmapped = [h for h in headers if h not in used_sources and str(h or "").strip()]
    return mapping, unmapped


def _clean(v):
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("none", "nan", "null", "n/a", "-") else s


def rows_to_records(rows):
    """rows: list of dicts (header -> value). Returns (records, columns_detected, unmapped)."""
    if not rows:
        return [], {}, []
    headers = list(rows[0].keys())
    mapping, unmapped = build_header_map(headers)
    records = []
    for i, row in enumerate(rows):
        rec = {k: "" for k in CANONICAL}
        rec["labels"] = []
        rec["extra"] = {}
        for canon, src in mapping.items():
            val = _clean(row.get(src))
            if canon == "labels":
                rec["labels"] = [t.strip() for t in re.split(r"[,;|]", val) if t.strip()]
            elif canon == "estimate":
                m = re.search(r"-?\d+(?:\.\d+)?", val)
                rec["estimate"] = (float(m.group()) if "." in m.group() else int(m.group())) if m else ""
            else:
                rec[canon] = val
        for src in unmapped:
            cv = _clean(row.get(src))
            if cv:
                rec["extra"][str(src)] = cv
        if not rec["id"]:
            rec["id"] = "ROW-%d" % (i + 1)
        # Drop fully empty rows (no title and no id beyond the synthetic one).
        if not rec["title"] and not any(rec[k] for k in ("description", "status", "priority")) and not rec["labels"]:
            if rec["id"].startswith("ROW-"):
                continue
        records.append(rec)
    return records, mapping, unmapped


# ----------------------------- format readers ------------------------------ #

def read_delimited(path, delimiter=None):
    with io.open(path, "r", encoding="utf-8-sig", newline="") as f:
        sample = f.read(4096)
        f.seek(0)
        if delimiter is None:
            try:
                delimiter = csv.Sniffer().sniff(sample, delimiters=",\t;|").delimiter
            except Exception:
                delimiter = "\t" if "\t" in sample.splitlines()[0] else ","
        reader = csv.DictReader(f, delimiter=delimiter)
        return [dict(r) for r in reader]


def read_json(path):
    with io.open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    # Accept a bare list, {"records":[...]}, {"cards":[...]} (re-import), or {"data":[...]}.
    if isinstance(data, dict):
        for key in ("records", "cards", "items", "data", "rows"):
            if isinstance(data.get(key), list):
                data = data[key]
                break
    if not isinstance(data, list):
        raise ValueError("JSON must be a list of objects or contain a list under records/cards/items/data.")
    rows = []
    for obj in data:
        if isinstance(obj, dict):
            rows.append({k: obj[k] for k in obj})
        else:
            rows.append({"title": str(obj)})
    return rows


def _ensure_openpyxl():
    try:
        import openpyxl  # noqa
        return
    except ImportError:
        pass
    sys.stderr.write("openpyxl not found; installing it for .xlsx support...\n")
    # Try progressively, to cope with PEP 668 'externally-managed' environments.
    attempts = [
        ["--quiet", "openpyxl"],
        ["--quiet", "--user", "openpyxl"],
        ["--quiet", "--user", "--break-system-packages", "openpyxl"],
    ]
    for flags in attempts:
        try:
            subprocess.run([sys.executable, "-m", "pip", "install"] + flags,
                           check=True, capture_output=True)
        except (subprocess.CalledProcessError, OSError):
            continue
        try:
            import openpyxl  # noqa
            return
        except ImportError:
            continue
    sys.exit(
        "Could not import or install openpyxl (needed for .xlsx).\n"
        "Fix it one of these ways:\n"
        "  • pip install openpyxl   (add --break-system-packages if your Python is 'externally managed')\n"
        "  • or re-export the sheet as CSV and pass that instead."
    )


def read_xlsx(path, sheet):
    _ensure_openpyxl()
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.worksheets[int(sheet)] if str(sheet).isdigit() else wb[sheet] if sheet else wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows_iter))
    except StopIteration:
        return []
    header = [str(h) if h is not None else ("col%d" % i) for i, h in enumerate(header)]
    rows = []
    for r in rows_iter:
        if r is None or all(c is None for c in r):
            continue
        rows.append({header[i]: (r[i] if i < len(r) else None) for i in range(len(header))})
    wb.close()
    return rows


def detect_format(path, fmt):
    if fmt != "auto":
        return fmt
    ext = os.path.splitext(path)[1].lower()
    return {".csv": "csv", ".tsv": "tsv", ".txt": "csv",
            ".json": "json", ".xlsx": "xlsx", ".xlsm": "xlsx"}.get(ext, "csv")


def main():
    ap = argparse.ArgumentParser(description="Normalize a list/spreadsheet into priority-board records.")
    ap.add_argument("--input", required=True, help="Path to csv/tsv/xlsx/json file")
    ap.add_argument("--format", default="auto", choices=["auto", "csv", "tsv", "xlsx", "json"])
    ap.add_argument("--sheet", default=None, help="Sheet name or 0-based index (xlsx only)")
    ap.add_argument("--out", default=None, help="Write JSON here instead of stdout")
    args = ap.parse_args()

    if not os.path.exists(args.input):
        sys.exit("Input file not found: %s" % args.input)

    fmt = detect_format(args.input, args.format)
    if fmt == "csv":
        rows = read_delimited(args.input, ",")
    elif fmt == "tsv":
        rows = read_delimited(args.input, "\t")
    elif fmt == "json":
        rows = read_json(args.input)
    elif fmt == "xlsx":
        rows = read_xlsx(args.input, args.sheet)
    else:
        sys.exit("Unsupported format: %s" % fmt)

    records, mapping, unmapped = rows_to_records(rows)
    out = {
        "source": os.path.basename(args.input),
        "format": fmt,
        "columns_detected": mapping,
        "unmapped": unmapped,
        "count": len(records),
        "records": records,
    }
    text = json.dumps(out, indent=2, ensure_ascii=False)
    if args.out:
        with io.open(args.out, "w", encoding="utf-8") as f:
            f.write(text)
        sys.stderr.write("Wrote %d records -> %s\n" % (len(records), args.out))
    else:
        print(text)


if __name__ == "__main__":
    main()
